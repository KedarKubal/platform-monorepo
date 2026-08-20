"""
Builds tracking-plan.xlsx for the Northline Goods Adobe Analytics demo project.
Run: python3 build_tracking_plan.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name=FONT_NAME, size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F3864")
SUBTITLE_FONT = Font(name=FONT_NAME, italic=True, size=9, color="595959")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

wb = openpyxl.Workbook()
wb.remove(wb.active)


def add_sheet(name, title, subtitle, headers, rows, col_widths):
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = subtitle
    ws["A2"].font = SUBTITLE_FONT
    header_row = 4
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for r, row in enumerate(rows, start=header_row + 1):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = WRAP
    for c, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row}"
    ws.row_dimensions[1].height = 20
    return ws


# ---------------- Events tab ----------------
events_headers = ["Var #", "Event Name", "Type", "Fires On", "Serialization", "Notes"]
events_rows = [
    ["event1", "Product View", "Counter", "PDP load", "Page load", "Fires once per PDP view"],
    ["event2", "Add to Cart", "Counter", "'Add to Cart' click", "Link tracking (s.tl)", "Linked to products string"],
    ["event3", "Remove from Cart", "Counter", "'Remove' click on cart page", "Link tracking", "Linked to products string"],
    ["event4", "Checkout Initiated", "Counter", "'Checkout' click from cart", "Link tracking", "Cart -> checkout entry point"],
    ["event5", "Checkout Step Complete", "Counter", "'Continue' click on each checkout step", "Link tracking", "Differentiate via eVar9"],
    ["event6", "Promo Code Applied - Success", "Counter", "Valid promo code submitted", "Link tracking", ""],
    ["event7", "Promo Code Applied - Failure", "Counter", "Invalid promo code submitted", "Link tracking", ""],
    ["event8", "Internal Search Performed", "Counter", "Search submitted", "Link tracking", "Paired with eVar6 / eVar7"],
    ["event9", "Purchase", "Counter", "Order confirmation page load", "Page load", "Should reconcile to Orders metric"],
    ["purchase", "Revenue / Units / Orders", "Currency / Counter", "Order confirmation page load", "Page load, in products string", "Standard commerce metrics"],
]
add_sheet(
    "Events", "Success Events (Metrics)",
    "Northline Goods demo - Adobe Analytics tracking plan",
    events_headers, events_rows, [10, 28, 16, 34, 24, 34],
)

# ---------------- eVars tab ----------------
evar_headers = ["Var #", "Name", "Description", "Value Example", "Allocation", "Expiration", "Data Layer Source"]
evar_rows = [
    ["eVar1", "Page Name", "Canonical page/template name", "PDP: Trail Runner Jacket", "Most Recent", "Visit", "digitalData.page.pageName"],
    ["eVar2", "Site Section", "Top-level nav section", "Outerwear", "Most Recent", "Visit", "digitalData.page.section"],
    ["eVar3", "Product Name", "Product viewed/acted on", "Trail Runner Jacket", "Most Recent", "Visit", "digitalData.product[0].name"],
    ["eVar4", "Product SKU", "Unique product identifier", "NLG-JCK-1042", "Most Recent", "Visit", "digitalData.product[0].sku"],
    ["eVar5", "Product Category", "Product's category/subcategory", "Outerwear/Jackets", "Most Recent", "Visit", "digitalData.product[0].category"],
    ["eVar6", "Internal Search Term", "Raw search query, lowercased", "waterproof jacket", "Most Recent", "Visit", "digitalData.search.term"],
    ["eVar7", "Internal Search Results Count", "# results returned", "14", "Most Recent", "Event (hit)", "digitalData.search.resultsCount"],
    ["eVar8", "Cart ID", "Session-scoped cart identifier", "cart_8f3ac1", "Most Recent", "Visit", "digitalData.cart.id"],
    ["eVar9", "Checkout Step Name", "Current checkout step", "shipping / payment / review", "Most Recent", "Event (hit)", "digitalData.checkout.stepName"],
    ["eVar10", "Order ID", "Completed order identifier", "NLG-100234", "Most Recent", "Event (hit)", "digitalData.order.orderId"],
    ["eVar11", "Promo Code", "Promo code entered", "WELCOME10", "Most Recent", "Event (hit)", "digitalData.promo.code"],
    ["eVar12", "PLP Sort/Filter Selection", "Active sort or filter facet", "sort:price_asc", "Most Recent", "Event (hit)", "digitalData.plp.activeFilter"],
    ["eVar13", "Marketing Channel", "Derived from utm_source/utm_medium", "paid_social", "Original Value", "Visit", "Query param parsed on landing"],
    ["eVar14", "Visitor Type", "New vs. returning (1st-party cookie check)", "new / returning", "Most Recent", "Visit", "Set in Launch rule via cookie"],
    ["eVar15", "Error Flag", "Marks 404/error pages", "404", "Most Recent", "Event (hit)", "digitalData.page.errorCode"],
]
add_sheet(
    "eVars", "Conversion Variables (eVars)",
    "Northline Goods demo - Adobe Analytics tracking plan",
    evar_headers, evar_rows, [8, 26, 34, 26, 14, 12, 32],
)

# ---------------- Props tab ----------------
prop_headers = ["Var #", "Name", "Mirrors", "Purpose"]
prop_rows = [
    ["prop1", "Page Name", "eVar1", "Enables Pathing/Fallout on page name without persistence"],
    ["prop2", "Site Section", "eVar2", "Section-level pathing"],
    ["prop3", "Product Name", "eVar3", "Product pathing - which products get viewed before abandon"],
    ["prop9", "Checkout Step Name", "eVar9", "Checkout pathing/fallout report"],
    ["prop12", "PLP Sort/Filter Selection", "eVar12", "Fallout on filter usage before exit"],
]
add_sheet(
    "Props", "Traffic Variables (Props)",
    "Northline Goods demo - Adobe Analytics tracking plan",
    prop_headers, prop_rows, [10, 26, 12, 44],
)

# ---------------- Products String tab ----------------
prod_headers = ["Scenario", "s.products Syntax", "Example"]
prod_rows = [
    ["Add to Cart", "Category;Product Name;Quantity;Price;event2=1;eVar4=SKU",
     "Outerwear;Trail Runner Jacket;1;129.00;event2=1;eVar4=NLG-JCK-1042"],
    ["Remove from Cart", "Category;Product Name;Quantity;Price;event3=1;eVar4=SKU",
     "Outerwear;Trail Runner Jacket;1;129.00;event3=1;eVar4=NLG-JCK-1042"],
    ["Purchase (per line item)", "Category;Product Name;Quantity;Price;event9=1,purchase;eVar4=SKU",
     "Outerwear;Trail Runner Jacket;1;129.00;event9=1,purchase;eVar4=NLG-JCK-1042"],
]
add_sheet(
    "Products String", "s.products Convention",
    "Northline Goods demo - Adobe Analytics tracking plan",
    prod_headers, prod_rows, [26, 46, 50],
)

# ---------------- Page Map tab ----------------
page_headers = ["Page", "Fires", "Variables Set"]
page_rows = [
    ["Home", "Page View", "eVar1, eVar2, eVar13 (first visit), eVar14"],
    ["PLP / Search Results", "Page View, event8 (if from search)", "eVar1, eVar2, eVar6, eVar7, prop1, prop2"],
    ["PLP - filter/sort applied", "Link tracking (no new page view)", "eVar12, prop12"],
    ["PDP", "Page View, event1", "eVar1-eVar5, prop1-prop3"],
    ["Add to Cart (any page)", "Link tracking, event2", "eVar4, eVar8, products string"],
    ["Cart page", "Page View", "eVar1, eVar8"],
    ["Remove from Cart", "Link tracking, event3", "eVar4, eVar8, products string"],
    ["Checkout - Shipping", "Page View, event4 (step 1 only), event5", "eVar1, eVar8, eVar9=shipping, prop9"],
    ["Checkout - Payment", "Page View, event5", "eVar9=payment, prop9"],
    ["Promo code field", "Link tracking, event6 or event7", "eVar11"],
    ["Checkout - Review", "Page View, event5", "eVar9=review, prop9"],
    ["Order Confirmation", "Page View, event9, purchase (products string)", "eVar1, eVar10, products string w/ per-item event9/purchase"],
    ["404 / Error page", "Page View", "eVar15"],
]
add_sheet(
    "Page Map", "Page -> Event/Variable Map",
    "Northline Goods demo - Adobe Analytics tracking plan",
    page_headers, page_rows, [26, 40, 50],
)

wb.save("tracking-plan.xlsx")
print("Saved tracking-plan.xlsx")
